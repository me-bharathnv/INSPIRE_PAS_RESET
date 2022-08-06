from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
import time
import openpyxl
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

path = "E:/inspire.xlsx" #Excel file path

wb = openpyxl.load_workbook(path)

sheet = wb.active

browser = webdriver.Chrome(executable_path="C:/Users/USER/OneDrive/Desktop/chromedriver.exe")

browser.get("https://www.inspireawards-dst.gov.in/UserC/login.aspx?to=2")
time.sleep(5)
browser.find_element(By.ID, "ctl00_ContentPlaceHolder1_txtUserId").send_keys("tumkur77")
browser.find_element(By.ID, "ctl00_ContentPlaceHolder1_txtPassword").send_keys("Diet@t22")
time.sleep(5)
browser.find_element(By.ID, "ctl00_ContentPlaceHolder1_txtCaptchaText").send_keys(input(str()))
time.sleep(10)
browser.find_element(By.ID, "ctl00_ContentPlaceHolder1_imgLogin").click()
time.sleep(5)

browser.find_element(By.XPATH, "//*[@id='wrapper']/div/section/div/div/div/div/div[1]/div[2]/div[2]/div[1]/div[7]/div[1]/div/h5").click()
time.sleep(5)

for i in range(6, sheet.max_row+1):
    app_code = sheet.cell(row=i, column=4)
    browser.find_element(By.ID, "ctl00_ContentPlaceHolder1_txtAppCode").send_keys(app_code.value)
    time.sleep(2)
    browser.find_element(By.ID, "ctl00_lblRegion").click()
    time.sleep(2)
    browser.find_element(By.ID, "ctl00_ContentPlaceHolder1_txtNewPass").send_keys("kglins@22")
    browser.find_element(By.ID, "ctl00_ContentPlaceHolder1_txtRePass").send_keys("kglins@22")
    time.sleep(2)
    browser.find_element(By.ID, "ctl00_ContentPlaceHolder1_btnChangePassword").click()
    WebDriverWait(browser, 5).until(EC.alert_is_present())
    browser.switch_to.alert.accept()
    time.sleep(2)
    sc_name = sheet.cell(row=i, column=2)
    print(f"{sc_name.value} school password changed successfully")
